# _*_coding : UTF-8 _*_
# @Time : 2026/2/7 17:35
# @Author : Murchey
# @File : getStandardData
# @Project : python
from openpyxl import load_workbook, Workbook
from pathlib import Path
import sys

def col_letter_to_index(col):
    """将列字母转换为整数索引（从0开始）"""
    col = col.upper()
    index = 0
    for char in col:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1

def getValuableData(filePath, beginRow, nameCol, moneyCol):
    """
    从Excel文件中提取数据
    参数：
        filePath: Excel文件路径
        beginRow: 表头所在的行数
        nameCol: 姓名栏目（如"A"、"B"等），作为字典的键
        moneyCol: 金额栏目（如"A"、"B"等），作为字典的值
    返回值：
        包含数据的字典 {姓名: 金额}
    """
    try:
        # 加载工作簿
        wb = load_workbook(filePath)
        ws = wb.active
        
        # 将列字母转换为索引
        name_col_index = col_letter_to_index(nameCol)
        money_col_index = col_letter_to_index(moneyCol)
        
        # 初始化结果字典
        data_dict = {}
        
        # 从beginRow + 1行开始读取数据（因为表头在beginRow行）
        for row in ws.iter_rows(min_row=beginRow + 1, values_only=True):
            # 检查行是否为空
            if not row:
                continue
            
            # 检查姓名列是否有值
            if len(row) > name_col_index and row[name_col_index] is not None:
                name = str(row[name_col_index]).strip()
                if name:
                    # 获取金额值
                    money = None
                    if len(row) > money_col_index and row[money_col_index] is not None:
                        # 尝试将金额转换为数字
                        try:
                            money = float(row[money_col_index])
                        except (ValueError, TypeError):
                            # 如果转换失败，尝试处理字符串格式的金额
                            try:
                                money_str = str(row[money_col_index]).strip()
                                # 移除货币符号和千位分隔符
                                money_str = money_str.replace(',', '').replace('￥', '').replace('$', '').replace('¥', '')
                                money = float(money_str)
                            except (ValueError, TypeError):
                                money = None
                    
                    # 将数据添加到字典
                    data_dict[name] = money
        
        return data_dict
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        return {}

def saveFile(dataDict, filePath):
    """
    将字典数据保存为Excel文件
    参数：
        dataDict: 要储存为xlsx的字典 {姓名: 金额}
        filePath: 要储存为xlsx的文件位置
    功能：
        能在软件被打包情况下也能正常使用
    """
    try:
        # 确保文件路径的目录存在
        file_path = Path(filePath)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        
        # 设置表头
        ws.cell(row=1, column=1, value="姓名")
        ws.cell(row=1, column=2, value="金额")
        
        # 填充数据
        row_index = 2
        for name, money in dataDict.items():
            ws.cell(row=row_index, column=1, value=name)
            ws.cell(row=row_index, column=2, value=money)
            row_index += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        # 保存文件
        wb.save(filePath)
        return True
    except Exception as e:
        print(f"保存Excel文件失败: {e}")
        return False