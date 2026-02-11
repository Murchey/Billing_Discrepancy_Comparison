# _*_coding : UTF-8 _*_
# @Time : 2026/2/6 19:45
# @Author : Murchey
# @File : main
# @Project : python
from collections import defaultdict
from openpyxl import load_workbook, Workbook
import os
from pathlib import Path

def col_letter_to_index(col):
    """
    将列字母（如'A'、'B'、'AA'等）转换为对应的整数索引（从0开始）
    支持大写和小写字母
    """
    col = col.upper()  # 转换为大写
    index = 0
    for char in col:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1  # 转换为从0开始的索引

def mainFunc(filename,newFileName,nameCol,moneyCol):
    print("读取:", filename)
    print("写出:", newFileName)
    # 确保 filename 是字符串
    from pathlib import Path
    if isinstance(filename, Path):
        filename = str(filename)
    
    # 处理列参数，将字母转换为整数索引
    if isinstance(nameCol, str):
        nameCol = col_letter_to_index(nameCol)
    if isinstance(moneyCol, str):
        moneyCol = col_letter_to_index(moneyCol)
    
    print(f"姓名列索引: {nameCol}, 金额列索引: {moneyCol}")
    
    try:
        wb = load_workbook(filename=filename, read_only=False, data_only=True)
        print("成功加载工作簿")
        ws = wb.active
        print("成功获取活动工作表")
    except Exception as e:
        print(f"加载工作簿失败: {e}")
        raise

    # print("当前检测到的工作表有：")
    # print(wb.sheetnames)

    stu_pay= defaultdict(float)
    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):
        # 检查行是否为空
        if not row:
            continue
        
        # 检查姓名列索引是否有效
        if nameCol >= len(row):
            print(f"警告：姓名列索引 {nameCol} 超出了行的长度 {len(row)}，跳过此行")
            continue
        
        # 检查金额列索引是否有效
        if moneyCol >= len(row):
            print(f"警告：金额列索引 {moneyCol} 超出了行的长度 {len(row)}，跳过此行")
            continue
        
        name = row[nameCol]
        money = row[moneyCol]
        
        # 检查姓名是否为空
        if name is None:
            continue
        
        # 检查金额是否为空
        if money is None:
            money = 0
        
        try:
            stu_pay[name] += float(money)
        except (ValueError, TypeError):
            print(f"警告：无法将金额 '{money}' 转换为数字，跳过此行")
            continue

    # print("提取姓名和个人缴费金额：")
    # print(dict(stu_pay))

    # 确保目录存在
    import sys
    from pathlib import Path
    
    # 获取应用程序的根目录
    import multipleFiles as mf
    app_dir = mf.get_application_dir()
    
    # 构建绝对路径
    save_dir = app_dir / newFileName
    save_dir.mkdir(parents=True, exist_ok=True)
    
    # 获取原始文件名（不含路径）
    base_filename = os.path.basename(filename)
    # 移除扩展名
    base_name = os.path.splitext(base_filename)[0]
    
    wb_res = Workbook()
    ws_res = wb_res.active
    ws_res.append(["姓名","缴费金额"])
    for stu,money in stu_pay.items():
        ws_res.append([stu,money])
    # 保存到指定目录，使用原始文件名作为基础
    save_path = save_dir / f"{base_name}_new.xlsx"
    wb_res.save(filename=save_path)
    print(f"文件已保存到: {save_path}")
    wb_res.close()