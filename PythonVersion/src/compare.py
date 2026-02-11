# _*_coding : UTF-8 _*_
# @Time : 2026/2/6 20:30
# @Author : Murchey
# @File : compare
# @Project : python
from itertools import zip_longest
from openpyxl import load_workbook, Workbook
import os
from pathlib import Path

def sheet_dict(ws, key_col=1, min_row=2):
    """
    把 ws 变成 dict： key(指定列) → 整行 tuple（values_only）
    返回：{key: (v1, v2, ..., vn), ...}
    """
    d = {}
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if row[key_col-1] is None:        # key 列空就跳过
            continue
        d[row[key_col-1]] = row
    return d


def diff_sheets(ws1, ws2, key_col=1, min_row=2):
    d1 = sheet_dict(ws1, key_col, min_row)
    d2 = sheet_dict(ws2, key_col, min_row)

    # 1. 行级差异
    only_in_1 = set(d1) - set(d2)
    only_in_2 = set(d2) - set(d1)
    common = set(d1) & set(d2)

    # 2. 单元格级差异
    cell_diff = []  # [(key, col_name, val1, val2), ...]
    headers = [ws1.cell(row=1, column=c).value
               for c in range(1, ws1.max_column + 1)]

    for k in common:
        row1, row2 = d1[k], d2[k]
        for col_idx, (v1, v2) in enumerate(zip_longest(row1, row2), 1):
            if v1 != v2:  # 不等就记录
                cell_diff.append(
                    (k, headers[col_idx - 1] if col_idx <= len(headers) else f'Col{col_idx}', v1, v2))

    return only_in_1, only_in_2, cell_diff

def compare_and_save(file1, file2, output_dir="比对结果"):
    """比较两个表格文件并将结果保存为Excel文件"""
    import os
    from pathlib import Path
    from openpyxl import Workbook, load_workbook
    import datetime
    
    # 获取应用程序的根目录，无论是脚本还是打包后的exe
    import multipleFiles as mf
    app_dir = mf.get_application_dir()
    
    # 构建完整的输出目录路径
    output_dir_path = app_dir / output_dir
    
    # 确保输出目录存在
    if not output_dir_path.exists():
        output_dir_path.mkdir(parents=True, exist_ok=True)
    
    # 确保 file1 和 file2 是字符串
    if isinstance(file1, Path):
        file1 = str(file1)
    if isinstance(file2, Path):
        file2 = str(file2)
    
    # 加载工作簿
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)
    
    # 获取文件名
    wb1_name = os.path.basename(file1)
    wb2_name = os.path.basename(file2)
    
    # 获取待比较文件的基础名称（不含扩展名）
    base_name = os.path.splitext(wb2_name)[0]
    
    # 比较表格
    only1, only2, cells = diff_sheets(wb1.active, wb2.active, key_col=1)  # A列当主键
    
    # 构建输出文件名，使用当前时间作为前缀
    current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"{current_time}_比对结果.xlsx"
    output_path = output_dir_path / output_filename
    
    # 检查文件是否已存在
    if output_path.exists():
        # 打开现有文件
        wb_output = load_workbook(output_path)
    else:
        # 创建新的工作簿
        wb_output = Workbook()
        # 删除默认的Sheet1
        if 'Sheet' in wb_output.sheetnames:
            wb_output.remove(wb_output['Sheet'])
    
    # 创建新的sheet，使用比对文件的名称，去掉_new后缀，然后添加_比对结果后缀
    sheet_name = base_name
    if sheet_name.endswith("_new"):
        sheet_name = sheet_name[:-4]  # 去掉 "_new"（4个字符）
    sheet_name = f"{sheet_name}_比对结果"
    # 如果sheet名已存在，添加数字后缀
    suffix = 1
    original_sheet_name = sheet_name
    while sheet_name in wb_output.sheetnames:
        sheet_name = f"{original_sheet_name}_{suffix}"
        suffix += 1
    
    # 创建新的sheet
    ws_output = wb_output.create_sheet(title=sheet_name)
    
    # 设置表头，字体大小为26号
    from openpyxl.styles import Font
    
    ws_output['A1'] = f'仅（标准文件）存在的行（{wb1_name}）'
    ws_output['B1'] = f'仅（待核对文件）存在的行（{wb2_name}）'
    ws_output['C1'] = '差异单元格（key, 列头, 标准文件值, 待核对文件值）'
    
    # 设置表头字体大小为26号
    font = Font(size=13, bold=True)
    ws_output['A1'].font = font
    ws_output['B1'].font = font
    ws_output['C1'].font = font
    
    # 写入数据
    max_rows = max(len(only1), len(only2), len(cells))
    
    # 写入第一列：仅wb1存在的行
    for i, k in enumerate(only1, 2):
        ws_output[f'A{i}'] = k
    
    # 写入第二列：仅wb2存在的行
    for i, k in enumerate(only2, 2):
        ws_output[f'B{i}'] = k
    
    # 写入第三列：差异单元格
    for i, (key, col, v1, v2) in enumerate(cells, 2):
        ws_output[f'C{i}'] = f"{key}, {col}, {v1}, {v2}"
    
    # 调整列宽，确保能容纳表头内容
    ws_output.column_dimensions['A'].width = 70
    ws_output.column_dimensions['B'].width = 70
    ws_output.column_dimensions['C'].width = 70
    
    # 保存为Excel文件
    wb_output.save(output_path)
    
    print(f"比对结果已保存到: {output_path}")
    return output_path

# 示例调用（如果直接运行此文件）
if __name__ == "__main__":
    import sys
    # 获取应用程序的根目录，无论是脚本还是打包后的exe
    if getattr(sys, 'frozen', False):
        # 打包后的exe模式
        app_dir = Path(sys.executable).resolve().parent
    else:
        # 脚本模式
        app_dir = Path(__file__).resolve().parent
    
    # 示例文件路径
    std_file = app_dir / "标准表格" / "TheRight.xlsx"
    test_file = app_dir / "已提取数据" / "ClassOne_new.xlsx"
    
    # 检查文件是否存在
    if std_file.exists() and test_file.exists():
        compare_and_save(std_file, test_file)
    else:
        print("示例文件不存在，跳过测试")
        print(f"标准表格文件: {std_file} {'存在' if std_file.exists() else '不存在'}")
        print(f"已提取数据文件: {test_file} {'存在' if test_file.exists() else '不存在'}")
